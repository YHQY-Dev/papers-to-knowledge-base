from pathlib import Path

from openpyxl import load_workbook

from papers_library_pipeline.export_excel import EXCEL_COLUMNS, export_literature_excel


def test_export_literature_excel_writes_required_columns(tmp_path: Path):
    items = [
        {
            "local_id": 1001,
            "title": "Example Review",
            "type": "review",
            "doi": "10.1000/xyz",
            "isbn": None,
            "year": 2020,
            "script_score": 80,
            "ai_score": 90,
            "accepted": True,
            "decision": "download",
            "reason": "core methods review",
            "pdf_status": "downloaded",
            "pdf_path": "myfield-pdf/1001.pdf",
            "notes": "",
        }
    ]
    out = tmp_path / "literature.xlsx"
    export_literature_excel(items, out)
    assert out.is_file()
    wb = load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == EXCEL_COLUMNS
    assert ws[2][0].value == 1001
    assert ws[2][1].value == "Example Review"
    assert ws[2][8].value in ("yes", True, "Y", "是")


def test_export_maps_selected_to_accepted_column(tmp_path: Path):
    items = [{"local_id": 1, "title": "Sel", "selected": True}]
    out = tmp_path / "literature.xlsx"
    export_literature_excel(items, out)
    ws = load_workbook(out).active
    assert ws[2][8].value == "yes"


def test_export_derives_pdf_status_from_disk(tmp_path: Path):
    pdf_dir = tmp_path / "pdf"
    pdf_dir.mkdir()
    (pdf_dir / "1001.Has_PDF.pdf").write_bytes(b"%PDF-1.4")
    items = [
        {"local_id": 1001, "title": "Has PDF", "accepted": True},
        {"local_id": 1002, "title": "Missing", "accepted": True},
    ]
    out = tmp_path / "literature.xlsx"
    export_literature_excel(items, out, pdf_dir=pdf_dir)
    ws = load_workbook(out).active
    # columns: ... accepted=8, decision=9, reason=10, pdf_status=11, pdf_path=12
    assert ws[2][11].value == "downloaded"
    assert str(ws[2][12].value).endswith("1001.Has_PDF.pdf")
    assert ws[3][11].value == "not_attempted"
    assert ws[3][12].value in ("", None)
